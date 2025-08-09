#!/usr/bin/env python3
"""
AI Agent System for Data Quality and Maintenance Automation
Python-based implementation for integration with Flask application
"""

import asyncio
import json
import logging
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
from dataclasses import dataclass
import uuid
import os
import sqlite3
from pathlib import Path

logger = logging.getLogger(__name__)

@dataclass
class AgentConfig:
    """Configuration for AI agents"""
    validation_rule_enabled: bool = True
    ab_testing_enabled: bool = True
    missing_data_enabled: bool = True
    script_generation_enabled: bool = True
    schedule_interval_minutes: int = 30
    max_concurrent_agents: int = 4

class ValidationRule:
    """Validation rule data structure"""
    def __init__(self, name: str, description: str, sql_condition: str, severity: str = "Medium"):
        self.id = str(uuid.uuid4())
        self.name = name
        self.description = description
        self.sql_condition = sql_condition
        self.severity = severity
        self.category = "DataQuality"
        self.created_by = "AI Agent"
        self.created_at = datetime.utcnow()
        self.is_active = True
        self.performance_metrics = {
            "success_rate": 0.0,
            "false_positive_rate": 0.0,
            "execution_time_ms": 0,
            "last_executed": datetime.utcnow()
        }

class DataPattern:
    """Data pattern for rule generation"""
    def __init__(self, pattern_type: str, field_name: str, confidence: float, description: str):
        self.id = str(uuid.uuid4())
        self.pattern_type = pattern_type
        self.field_name = field_name
        self.confidence = confidence
        self.pattern_description = description

class ABTest:
    """A/B test data structure"""
    def __init__(self, rule_a: ValidationRule, rule_b: ValidationRule):
        self.id = str(uuid.uuid4())
        self.name = f"A/B Test: {rule_a.name} vs {rule_b.name}"
        self.rule_a = rule_a
        self.rule_b = rule_b
        self.start_date = datetime.utcnow()
        self.end_date = None
        self.status = "Running"
        self.metrics = {
            "rule_a_performance": 0.0,
            "rule_b_performance": 0.0,
            "statistical_significance": 0.0,
            "sample_size": 0,
            "confidence_interval": (0.0, 0.0)
        }

class MissingDataItem:
    """Missing data item structure"""
    def __init__(self, entity_type: str, entity_id: str, missing_fields: List[str], priority: str = "Medium"):
        self.id = str(uuid.uuid4())
        self.entity_type = entity_type
        self.entity_id = entity_id
        self.missing_fields = missing_fields
        self.priority = priority
        self.estimated_effort = 30  # minutes
        self.business_impact = "Moderate"

class ExcelScript:
    """Excel script data structure"""
    def __init__(self, name: str, description: str, entity_type: str):
        self.id = str(uuid.uuid4())
        self.name = name
        self.description = description
        self.template_path = f"/app/templates/{entity_type.lower()}_template.xlsx"
        self.validation_rules = []
        self.data_mapping = {
            "source_fields": [],
            "target_fields": [],
            "transformation_rules": []
        }
        self.status = "Draft"
        self.created_at = datetime.utcnow()
        self.approved_at = None

class Task:
    """Task data structure"""
    def __init__(self, title: str, description: str, priority: str = "Medium"):
        self.id = str(uuid.uuid4())
        self.title = title
        self.description = description
        self.priority = priority
        self.status = "Open"
        self.assigned_to = None
        self.due_date = datetime.utcnow() + timedelta(days=7)
        self.created_at = datetime.utcnow()
        self.entity_type = "Equipment"
        self.entity_id = ""

class BaseAgent:
    """Base class for all AI agents"""
    def __init__(self, name: str):
        self.name = name
        self.is_active = True
        self.last_execution = None
        self.metrics = {}
        self.error_count = 0

    async def execute(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        """Execute the agent with input data"""
        try:
            self.last_execution = datetime.utcnow()
            result = await self._execute_impl(input_data)
            return {
                "success": True,
                "data": result,
                "message": f"{self.name} executed successfully",
                "timestamp": datetime.utcnow().isoformat()
            }
        except Exception as e:
            self.error_count += 1
            logger.error(f"Error in {self.name}: {e}")
            return {
                "success": False,
                "data": {},
                "message": str(e),
                "timestamp": datetime.utcnow().isoformat()
            }

    async def get_status(self) -> Dict[str, Any]:
        """Get agent status"""
        return {
            "is_active": self.is_active,
            "last_execution": self.last_execution.isoformat() if self.last_execution else None,
            "metrics": self.metrics,
            "error_count": self.error_count
        }

    async def _execute_impl(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        """Implementation specific to each agent - to be overridden"""
        raise NotImplementedError

class ValidationRuleAgent(BaseAgent):
    """Agent for creating and optimizing validation rules"""
    
    def __init__(self):
        super().__init__("ValidationRuleAgent")
        self.rule_repository = get_rule_repository()

    async def _execute_impl(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        dataset = input_data.get("dataset", "equipment_data")
        
        # Analyze data patterns
        patterns = await self._analyze_data_patterns(dataset)
        
        # Generate rules from patterns
        rules = await self._generate_rules(patterns)
        
        # Save rules to repository
        for rule in rules:
            await self.rule_repository.save_rule(rule)
        
        return {
            "rules_generated": len(rules),
            "rules": [{"id": rule.id, "name": rule.name, "description": rule.description} for rule in rules]
        }

    async def _analyze_data_patterns(self, dataset: str) -> List[DataPattern]:
        """Analyze data patterns for rule generation"""
        # Mock pattern analysis - in real implementation, this would use ML models
        patterns = [
            DataPattern("Completeness", "equipment_name", 0.85, "Missing equipment names detected"),
            DataPattern("Anomaly", "maintenance_date", 0.92, "Invalid maintenance dates detected"),
            DataPattern("Completeness", "manufacturer", 0.78, "Missing manufacturer information"),
            DataPattern("Anomaly", "serial_number", 0.88, "Invalid serial number format")
        ]
        return patterns

    async def _generate_rules(self, patterns: List[DataPattern]) -> List[ValidationRule]:
        """Generate validation rules from patterns"""
        rules = []
        
        for pattern in patterns:
            rule = ValidationRule(
                name=f"Auto-generated rule for {pattern.field_name}",
                description=pattern.pattern_description,
                sql_condition=self._generate_sql_condition(pattern),
                severity=self._determine_severity(pattern.confidence)
            )
            rules.append(rule)
        
        return rules

    def _generate_sql_condition(self, pattern: DataPattern) -> str:
        """Generate SQL condition for validation rule"""
        if pattern.pattern_type == "Completeness":
            return f"{pattern.field_name} IS NOT NULL"
        elif pattern.pattern_type == "Anomaly":
            return f"{pattern.field_name} IS NOT NULL AND {pattern.field_name} != ''"
        else:
            return f"{pattern.field_name} IS NOT NULL"

    def _determine_severity(self, confidence: float) -> str:
        """Determine rule severity based on confidence"""
        if confidence > 0.9:
            return "Critical"
        elif confidence > 0.7:
            return "High"
        elif confidence > 0.5:
            return "Medium"
        else:
            return "Low"

class ABTestingAgent(BaseAgent):
    """Agent for A/B testing validation rules"""
    
    def __init__(self):
        super().__init__("ABTestingAgent")
        self.test_repository = get_test_repository()

    async def _execute_impl(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        rule_a_data = input_data.get("rule_a")
        rule_b_data = input_data.get("rule_b")
        
        if not rule_a_data or not rule_b_data:
            raise ValueError("Both rule_a and rule_b must be provided")
        
        # Create validation rules from input data
        rule_a = ValidationRule(
            name=rule_a_data.get("name", "Rule A"),
            description=rule_a_data.get("description", ""),
            sql_condition=rule_a_data.get("sql_condition", "")
        )
        
        rule_b = ValidationRule(
            name=rule_b_data.get("name", "Rule B"),
            description=rule_b_data.get("description", ""),
            sql_condition=rule_b_data.get("sql_condition", "")
        )
        
        # Create A/B test
        test = ABTest(rule_a, rule_b)
        await self.test_repository.save_test(test)
        
        return {
            "test_id": test.id,
            "test_name": test.name,
            "status": test.status
        }

class MissingDataAgent(BaseAgent):
    """Agent for detecting missing data and creating tasks"""
    
    def __init__(self):
        super().__init__("MissingDataAgent")
        self.task_repository = get_task_repository()

    async def _execute_impl(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        # Scan for missing data
        missing_items = await self._scan_for_missing_data()
        
        # Create tasks for missing data
        tasks = await self._create_tasks(missing_items)
        
        return {
            "missing_items_found": len(missing_items),
            "tasks_created": len(tasks),
            "missing_items": [
                {
                    "entity_type": item.entity_type,
                    "entity_id": item.entity_id,
                    "missing_fields": item.missing_fields,
                    "priority": item.priority
                }
                for item in missing_items
            ]
        }

    async def _scan_for_missing_data(self) -> List[MissingDataItem]:
        """Scan system for missing data"""
        # Mock missing data scan - in real implementation, this would query the database
        missing_items = [
            MissingDataItem("Equipment", "EQ001", ["manufacturer", "model"], "High"),
            MissingDataItem("FunctionalLocation", "FL001", ["location_code"], "Medium"),
            MissingDataItem("Equipment", "EQ002", ["serial_number"], "High"),
            MissingDataItem("MaintenanceOrder", "MO001", ["priority"], "Medium")
        ]
        return missing_items

    async def _create_tasks(self, missing_items: List[MissingDataItem]) -> List[Task]:
        """Create tasks for missing data items"""
        tasks = []
        
        for item in missing_items:
            task = Task(
                title=f"Complete missing data for {item.entity_id}",
                description=f"Missing fields: {', '.join(item.missing_fields)}",
                priority=item.priority
            )
            task.entity_type = item.entity_type
            task.entity_id = item.entity_id
            
            await self.task_repository.save_task(task)
            tasks.append(task)
        
        return tasks

class TaskExecutionAgent(BaseAgent):
    """Agent for executing and managing tasks"""
    
    def __init__(self):
        super().__init__("TaskExecutionAgent")
        self.task_repository = get_task_repository()

    async def _execute_impl(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        """Execute task management operations"""
        operation = input_data.get("operation", "monitor")
        
        if operation == "monitor":
            return await self._monitor_tasks()
        elif operation == "execute":
            task_id = input_data.get("task_id")
            return await self._execute_task(task_id)
        elif operation == "update_status":
            task_id = input_data.get("task_id")
            status = input_data.get("status")
            completion = input_data.get("completion_percentage")
            return await self._update_task_status(task_id, status, completion)
        elif operation == "generate_report":
            return await self._generate_task_report()
        else:
            return {"error": f"Unknown operation: {operation}"}

    async def _monitor_tasks(self) -> Dict[str, Any]:
        """Monitor all tasks and their status"""
        try:
            all_tasks = await self.task_repository.get_tasks()
            overdue_tasks = await self.task_repository.get_overdue_tasks()
            
            # Calculate statistics
            total_tasks = len(all_tasks)
            completed_tasks = len([t for t in all_tasks if t['status'] == 'Completed'])
            in_progress_tasks = len([t for t in all_tasks if t['status'] == 'In Progress'])
            overdue_count = len(overdue_tasks)
            
            # Calculate average completion percentage
            completion_percentages = [t.get('completion_percentage', 0) for t in all_tasks if t.get('completion_percentage')]
            avg_completion = sum(completion_percentages) / len(completion_percentages) if completion_percentages else 0
            
            return {
                "total_tasks": total_tasks,
                "completed_tasks": completed_tasks,
                "in_progress_tasks": in_progress_tasks,
                "overdue_tasks": overdue_count,
                "average_completion": round(avg_completion, 2),
                "overdue_tasks_list": overdue_tasks[:5],  # Top 5 overdue
                "recent_tasks": all_tasks[:10]  # Recent 10 tasks
            }
        except Exception as e:
            logger.error(f"Error monitoring tasks: {e}")
            return {"error": str(e)}

    async def _execute_task(self, task_id: str) -> Dict[str, Any]:
        """Execute a specific task"""
        try:
            # Get task details
            tasks = await self.task_repository.get_tasks()
            task = next((t for t in tasks if t['id'] == task_id), None)
            
            if not task:
                return {"error": f"Task {task_id} not found"}
            
            # Simulate task execution based on task type
            task_type = task.get('task_type', 'Manual')
            automation_level = task.get('automation_level', 'None')
            
            if automation_level == 'Full':
                # Fully automated task
                await self.task_repository.update_task_status(task_id, 'Completed', 100.0)
                result = {"status": "completed", "automation": "full"}
            elif automation_level == 'Partial':
                # Partially automated task
                await self.task_repository.update_task_status(task_id, 'In Progress', 75.0)
                result = {"status": "in_progress", "automation": "partial", "completion": 75.0}
            else:
                # Manual task - mark as ready for human intervention
                await self.task_repository.update_task_status(task_id, 'Ready for Execution', 0.0)
                result = {"status": "ready_for_execution", "automation": "manual"}
            
            return {
                "task_id": task_id,
                "task_title": task['title'],
                "execution_result": result,
                "message": f"Task '{task['title']}' execution initiated"
            }
        except Exception as e:
            logger.error(f"Error executing task {task_id}: {e}")
            return {"error": str(e)}

    async def _update_task_status(self, task_id: str, status: str, completion_percentage: float = None) -> Dict[str, Any]:
        """Update task status and completion percentage"""
        try:
            await self.task_repository.update_task_status(task_id, status, completion_percentage)
            return {
                "task_id": task_id,
                "status": status,
                "completion_percentage": completion_percentage,
                "message": f"Task {task_id} status updated to {status}"
            }
        except Exception as e:
            logger.error(f"Error updating task {task_id}: {e}")
            return {"error": str(e)}

    async def _generate_task_report(self) -> Dict[str, Any]:
        """Generate comprehensive task management report"""
        try:
            all_tasks = await self.task_repository.get_tasks()
            overdue_tasks = await self.task_repository.get_overdue_tasks()
            
            # Calculate detailed statistics
            status_counts = {}
            priority_counts = {}
            automation_counts = {}
            
            for task in all_tasks:
                status = task.get('status', 'Unknown')
                priority = task.get('priority', 'Unknown')
                automation = task.get('automation_level', 'Unknown')
                
                status_counts[status] = status_counts.get(status, 0) + 1
                priority_counts[priority] = priority_counts.get(priority, 0) + 1
                automation_counts[automation] = automation_counts.get(automation, 0) + 1
            
            # Calculate efficiency metrics
            total_hours_estimated = sum(t.get('estimated_hours', 0) for t in all_tasks)
            total_hours_actual = sum(t.get('actual_hours', 0) for t in all_tasks)
            efficiency_ratio = total_hours_actual / total_hours_estimated if total_hours_estimated > 0 else 0
            
            return {
                "report_type": "task_management",
                "generated_at": datetime.utcnow().isoformat(),
                "summary": {
                    "total_tasks": len(all_tasks),
                    "overdue_tasks": len(overdue_tasks),
                    "efficiency_ratio": round(efficiency_ratio, 2),
                    "total_estimated_hours": round(total_hours_estimated, 2),
                    "total_actual_hours": round(total_hours_actual, 2)
                },
                "status_distribution": status_counts,
                "priority_distribution": priority_counts,
                "automation_distribution": automation_counts,
                "overdue_tasks": overdue_tasks,
                "recent_activities": all_tasks[:20]
            }
        except Exception as e:
            logger.error(f"Error generating task report: {e}")
            return {"error": str(e)}

class ScriptGenerationAgent(BaseAgent):
    """Agent for generating Excel scripts"""
    
    def __init__(self):
        super().__init__("ScriptGenerationAgent")

    async def _execute_impl(self, input_data: Dict[str, Any]) -> Dict[str, Any]:
        entity_type = input_data.get("entity_type", "Equipment")
        description = input_data.get("description", f"Update script for {entity_type}")
        
        # Generate Excel script
        script = ExcelScript(
            name=f"Update script for {entity_type}",
            description=description,
            entity_type=entity_type
        )
        
        # Generate template path
        template_path = await self._generate_template(entity_type)
        
        return {
            "script_id": script.id,
            "template_path": template_path,
            "script_name": script.name,
            "status": script.status
        }

    async def _generate_template(self, entity_type: str) -> str:
        """Generate Excel template for entity type"""
        template_dir = Path("/app/templates")
        template_dir.mkdir(exist_ok=True)
        
        template_path = template_dir / f"{entity_type.lower()}_template.xlsx"
        
        # Create actual Excel file with template structure
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{entity_type} Template"
            
            # Define headers based on entity type
            if entity_type.lower() == "equipment":
                headers = ["Equipment_ID", "Equipment_Name", "Manufacturer", "Model", "Serial_Number", "Location", "Status", "Last_Maintenance", "Next_Maintenance"]
            elif entity_type.lower() == "functionallocation":
                headers = ["Location_ID", "Location_Name", "Location_Code", "Parent_Location", "Equipment_Count", "Status", "Description"]
            elif entity_type.lower() == "maintenanceorder":
                headers = ["Order_ID", "Equipment_ID", "Order_Type", "Priority", "Description", "Planned_Date", "Actual_Date", "Status", "Assigned_To"]
            else:
                headers = ["ID", "Name", "Description", "Status", "Created_Date"]
            
            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Add headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add sample data row
            sample_data = ["SAMPLE_" + str(i+1) for i in range(len(headers))]
            for col, value in enumerate(sample_data, 1):
                ws.cell(row=2, column=col, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save the file
            wb.save(str(template_path))
            logger.info(f"Created Excel template: {template_path}")
            
        except ImportError:
            logger.warning("openpyxl not available, creating placeholder file")
            # Create a simple text file as fallback
            with open(template_path, 'w') as f:
                f.write(f"# {entity_type} Template\n")
                f.write("# This is a placeholder template\n")
                f.write("# Install openpyxl for proper Excel generation\n")
        
        return str(template_path)

class AIAgentManager:
    """Manages all AI agents and their workflows"""
    
    def __init__(self, config: AgentConfig):
        self.config = config
        self.agents = {}
        self.workflows = {}
        self.is_running = False
        
        # Initialize agents
        self._initialize_agents()

    def _initialize_agents(self):
        """Initialize all AI agents"""
        if self.config.validation_rule_enabled:
            self.agents['validation_rule'] = ValidationRuleAgent()
        
        if self.config.ab_testing_enabled:
            self.agents['ab_testing'] = ABTestingAgent()
        
        if self.config.missing_data_enabled:
            self.agents['missing_data'] = MissingDataAgent()
        
        if self.config.script_generation_enabled:
            self.agents['script_generation'] = ScriptGenerationAgent()
        
        # Add Task Execution Agent
        self.agents['task_execution'] = TaskExecutionAgent()
        
        logger.info(f"Initialized {len(self.agents)} AI agents")

    async def start_scheduled_workflows(self):
        """Start scheduled workflow execution"""
        self.is_running = True
        
        while self.is_running:
            try:
                await self._execute_scheduled_workflows()
                await asyncio.sleep(self.config.schedule_interval_minutes * 60)
            except Exception as e:
                logger.error(f"Error in scheduled workflows: {e}")
                await asyncio.sleep(60)  # Wait 1 minute before retrying

    async def _execute_scheduled_workflows(self):
        """Execute scheduled workflows"""
        workflows = [
            self._validation_rule_workflow,
            self._missing_data_workflow,
            self._ab_testing_workflow,
        ]
        
        # Execute workflows concurrently
        tasks = []
        for workflow in workflows:
            if self._should_execute_workflow(workflow.__name__):
                tasks.append(asyncio.create_task(workflow()))
        
        if tasks:
            results = await asyncio.gather(*tasks, return_exceptions=True)
            for i, result in enumerate(results):
                if isinstance(result, Exception):
                    logger.error(f"Workflow {i} failed: {result}")

    def _should_execute_workflow(self, workflow_name: str) -> bool:
        """Determine if a workflow should be executed based on configuration"""
        if workflow_name == "validation_rule_workflow":
            return self.config.validation_rule_enabled
        elif workflow_name == "ab_testing_workflow":
            return self.config.ab_testing_enabled
        elif workflow_name == "missing_data_workflow":
            return self.config.missing_data_enabled
        return True

    async def _validation_rule_workflow(self):
        """Validation rule creation workflow"""
        try:
            agent = self.agents['validation_rule']
            result = await agent.execute({"dataset": "equipment_data"})
            
            if result["success"]:
                logger.info(f"Validation rule workflow completed: {result['message']}")
                
                # Create A/B test for new rules if enabled
                if self.config.ab_testing_enabled:
                    await self._create_ab_test_for_rules(result["data"].get("rules", []))
            else:
                logger.error(f"Validation rule workflow failed: {result['message']}")
                
        except Exception as e:
            logger.error(f"Validation rule workflow error: {e}")

    async def _missing_data_workflow(self):
        """Missing data detection workflow"""
        try:
            agent = self.agents['missing_data']
            result = await agent.execute({})
            
            if result["success"]:
                logger.info(f"Missing data workflow completed: {result['message']}")
                
                # Generate scripts for missing data if enabled
                if self.config.script_generation_enabled:
                    await self._generate_scripts_for_missing_data(result["data"])
            else:
                logger.error(f"Missing data workflow failed: {result['message']}")
                
        except Exception as e:
            logger.error(f"Missing data workflow error: {e}")

    async def _ab_testing_workflow(self):
        """A/B testing workflow"""
        try:
            agent = self.agents['ab_testing']
            
            # Get active tests
            active_tests = await self._get_active_tests()
            
            for test in active_tests:
                # Monitor test performance
                metrics = await agent._monitor_test(test['id'])
                
                # Analyze results if test is complete
                if test['status'] == 'completed':
                    result = await agent._analyze_results(test['id'])
                    
                    if result.get('confidence', 0) > 0.9:  # Statistical significance threshold
                        await agent._promote_winner(test['id'])
                        logger.info(f"Promoted winning rule for test {test['id']}")
                
        except Exception as e:
            logger.error(f"A/B testing workflow error: {e}")

    async def _create_ab_test_for_rules(self, rules: List[Dict]):
        """Create A/B tests for newly generated rules"""
        try:
            agent = self.agents['ab_testing']
            
            for i in range(0, len(rules) - 1, 2):
                rule_a = rules[i]
                rule_b = rules[i + 1]
                
                test_input = {
                    'rule_a': rule_a,
                    'rule_b': rule_b,
                    'test_duration_days': 7
                }
                
                result = await agent.execute(test_input)
                
                if result["success"]:
                    logger.info(f"Created A/B test: {result['data'].get('test_name')}")
                    
        except Exception as e:
            logger.error(f"Error creating A/B tests: {e}")

    async def _generate_scripts_for_missing_data(self, data: Dict):
        """Generate Excel scripts for missing data"""
        try:
            agent = self.agents['script_generation']
            
            missing_items = data.get('missing_items_found', 0)
            
            if missing_items > 0:
                # Generate scripts for different entity types
                entity_types = ['Equipment', 'FunctionalLocation', 'MaintenanceOrder']
                
                for entity_type in entity_types:
                    script_input = {
                        'entity_type': entity_type,
                        'description': f'Update script for {entity_type} missing data',
                        'include_validations': True
                    }
                    
                    result = await agent.execute(script_input)
                    
                    if result["success"]:
                        logger.info(f"Generated script for {entity_type}: {result['data'].get('script_id')}")
                        
        except Exception as e:
            logger.error(f"Error generating scripts: {e}")

    async def _get_active_tests(self) -> List[Dict]:
        """Get active A/B tests from database"""
        # This would typically query the database
        # For now, return mock data
        return [
            {
                'id': 'test_001',
                'status': 'running',
                'start_date': datetime.now().isoformat()
            }
        ]

    async def execute_workflow(self, workflow_name: str, parameters: Dict = None) -> Dict:
        """Execute a specific workflow"""
        try:
            if workflow_name not in self.agents:
                raise ValueError(f"Unknown workflow: {workflow_name}")
            
            agent = self.agents[workflow_name]
            input_data = parameters or {}
            
            result = await agent.execute(input_data)
            
            return result
            
        except Exception as e:
            logger.error(f"Workflow execution error: {e}")
            return {
                'success': False,
                'message': str(e),
                'data': {},
                'timestamp': datetime.now().isoformat()
            }

    async def get_agent_status(self) -> Dict[str, Dict]:
        """Get status of all agents"""
        status = {}
        
        for name, agent in self.agents.items():
            try:
                agent_status = await agent.get_status()
                status[name] = agent_status
            except Exception as e:
                logger.error(f"Error getting status for agent {name}: {e}")
                status[name] = {
                    'is_active': False,
                    'last_execution': None,
                    'metrics': {},
                    'error_count': 0,
                    'error': str(e)
                }
        
        return status

    def stop(self):
        """Stop the agent manager"""
        self.is_running = False

# Repository classes for data persistence
# -----------------------------
# SQLite repositories (existing)
# -----------------------------

class RuleRepository:
    """Repository for validation rules"""
    
    def __init__(self):
        self.db_path = "/app/data/rules.db"
        self._init_db()

    def _init_db(self):
        """Initialize database"""
        os.makedirs(os.path.dirname(self.db_path), exist_ok=True)
        
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS validation_rules (
                    id TEXT PRIMARY KEY,
                    name TEXT NOT NULL,
                    description TEXT,
                    sql_condition TEXT NOT NULL,
                    severity TEXT,
                    category TEXT,
                    created_by TEXT,
                    created_at TEXT,
                    is_active BOOLEAN,
                    performance_metrics TEXT
                )
            """)

    async def save_rule(self, rule: ValidationRule):
        """Save validation rule to database"""
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("""
                INSERT OR REPLACE INTO validation_rules 
                (id, name, description, sql_condition, severity, category, created_by, created_at, is_active, performance_metrics)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                rule.id, rule.name, rule.description, rule.sql_condition,
                rule.severity, rule.category, rule.created_by,
                rule.created_at.isoformat(), rule.is_active,
                json.dumps(rule.performance_metrics)
            ))

class TestRepository:
    """Repository for A/B tests"""
    
    def __init__(self):
        self.db_path = "/app/data/tests.db"
        self._init_db()

    def _init_db(self):
        """Initialize database"""
        os.makedirs(os.path.dirname(self.db_path), exist_ok=True)
        
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS ab_tests (
                    id TEXT PRIMARY KEY,
                    name TEXT NOT NULL,
                    rule_a TEXT,
                    rule_b TEXT,
                    start_date TEXT,
                    end_date TEXT,
                    status TEXT,
                    metrics TEXT
                )
            """)

    async def save_test(self, test: ABTest):
        """Save A/B test to database"""
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("""
                INSERT OR REPLACE INTO ab_tests 
                (id, name, rule_a, rule_b, start_date, end_date, status, metrics)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                test.id, test.name,
                json.dumps({"id": test.rule_a.id, "name": test.rule_a.name}),
                json.dumps({"id": test.rule_b.id, "name": test.rule_b.name}),
                test.start_date.isoformat(),
                test.end_date.isoformat() if test.end_date else None,
                test.status,
                json.dumps(test.metrics)
            ))

class TaskRepository:
    """Repository for tasks"""
    
    def __init__(self):
        self.db_path = "/app/data/tasks.db"
        self._init_db()

    def _init_db(self):
        """Initialize database"""
        os.makedirs(os.path.dirname(self.db_path), exist_ok=True)
        
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS tasks (
                    id TEXT PRIMARY KEY,
                    title TEXT NOT NULL,
                    description TEXT,
                    priority TEXT,
                    status TEXT,
                    assigned_to TEXT,
                    due_date TEXT,
                    created_at TEXT,
                    entity_type TEXT,
                    entity_id TEXT,
                    task_type TEXT,
                    automation_level TEXT,
                    estimated_hours REAL,
                    actual_hours REAL,
                    completion_percentage REAL,
                    dependencies TEXT,
                    tags TEXT
                )
            """)

    async def save_task(self, task: Task):
        """Save task to database"""
        with sqlite3.connect(self.db_path) as conn:
            conn.execute("""
                INSERT OR REPLACE INTO tasks 
                (id, title, description, priority, status, assigned_to, due_date, created_at, entity_type, entity_id,
                 task_type, automation_level, estimated_hours, actual_hours, completion_percentage, dependencies, tags)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                task.id, task.title, task.description, task.priority,
                task.status, task.assigned_to,
                task.due_date.isoformat() if task.due_date else None,
                task.created_at.isoformat(),
                task.entity_type, task.entity_id,
                getattr(task, 'task_type', 'Manual'),
                getattr(task, 'automation_level', 'None'),
                getattr(task, 'estimated_hours', 0.0),
                getattr(task, 'actual_hours', 0.0),
                getattr(task, 'completion_percentage', 0.0),
                json.dumps(getattr(task, 'dependencies', [])),
                json.dumps(getattr(task, 'tags', []))
            ))

    async def get_tasks(self, status: str = None, priority: str = None) -> List[Dict]:
        """Get tasks with optional filtering"""
        with sqlite3.connect(self.db_path) as conn:
            query = "SELECT * FROM tasks WHERE 1=1"
            params = []
            
            if status:
                query += " AND status = ?"
                params.append(status)
            
            if priority:
                query += " AND priority = ?"
                params.append(priority)
            
            query += " ORDER BY created_at DESC"
            
            cursor = conn.execute(query, params)
            columns = [description[0] for description in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()]

    async def update_task_status(self, task_id: str, status: str, completion_percentage: float = None):
        """Update task status and completion percentage"""
        with sqlite3.connect(self.db_path) as conn:
            if completion_percentage is not None:
                conn.execute("""
                    UPDATE tasks SET status = ?, completion_percentage = ? WHERE id = ?
                """, (status, completion_percentage, task_id))
            else:
                conn.execute("""
                    UPDATE tasks SET status = ? WHERE id = ?
                """, (status, task_id))

    async def get_overdue_tasks(self) -> List[Dict]:
        """Get overdue tasks"""
        with sqlite3.connect(self.db_path) as conn:
            cursor = conn.execute("""
                SELECT * FROM tasks 
                WHERE due_date < ? AND status != 'Completed'
                ORDER BY due_date ASC
            """, (datetime.utcnow().isoformat(),))
            
            columns = [description[0] for description in cursor.description]
            return [dict(zip(columns, row)) for row in cursor.fetchall()] 


# ---------------------------------
# Postgres repositories (new)
# ---------------------------------

def use_postgres_storage() -> bool:
    storage = os.getenv("AI_AGENTS_STORAGE", "postgres").strip().lower()
    return storage in ("pg", "postgres", "postgresql")


def get_rule_repository():
    if use_postgres_storage():
        return PgRuleRepository()
    return RuleRepository()


def get_test_repository():
    if use_postgres_storage():
        return PgTestRepository()
    return TestRepository()


def get_task_repository():
    if use_postgres_storage():
        return PgTaskRepository()
    return TaskRepository()


try:
    # Import lazily so SQLite-only environments still work
    from .pg_repo import get_pg_connection
except Exception:
    # Fallback when running this module standalone
    from pg_repo import get_pg_connection  # type: ignore


class PgRuleRepository:
    def __init__(self):
        self._ensure_table()

    def _ensure_table(self):
        with get_pg_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS validation_rules (
                        id TEXT PRIMARY KEY,
                        name TEXT NOT NULL,
                        description TEXT,
                        sql_condition TEXT NOT NULL,
                        severity TEXT,
                        category TEXT,
                        created_by TEXT,
                        created_at TIMESTAMP,
                        is_active BOOLEAN,
                        performance_metrics JSONB
                    )
                    """
                )

    async def save_rule(self, rule: ValidationRule):
        with get_pg_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO validation_rules (
                        id, name, description, sql_condition, severity, category, created_by, created_at, is_active, performance_metrics
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (id) DO UPDATE SET
                        name = EXCLUDED.name,
                        description = EXCLUDED.description,
                        sql_condition = EXCLUDED.sql_condition,
                        severity = EXCLUDED.severity,
                        category = EXCLUDED.category,
                        created_by = EXCLUDED.created_by,
                        created_at = EXCLUDED.created_at,
                        is_active = EXCLUDED.is_active,
                        performance_metrics = EXCLUDED.performance_metrics
                    """,
                    (
                        rule.id,
                        rule.name,
                        rule.description,
                        rule.sql_condition,
                        rule.severity,
                        rule.category,
                        rule.created_by,
                        rule.created_at,
                        rule.is_active,
                        json.dumps(rule.performance_metrics),
                    ),
                )


class PgTestRepository:
    def __init__(self):
        self._ensure_table()

    def _ensure_table(self):
        with get_pg_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS ab_tests (
                        id TEXT PRIMARY KEY,
                        name TEXT NOT NULL,
                        rule_a JSONB,
                        rule_b JSONB,
                        start_date TIMESTAMP,
                        end_date TIMESTAMP,
                        status TEXT,
                        metrics JSONB
                    )
                    """
                )

    async def save_test(self, test: ABTest):
        with get_pg_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO ab_tests (
                        id, name, rule_a, rule_b, start_date, end_date, status, metrics
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (id) DO UPDATE SET
                        name = EXCLUDED.name,
                        rule_a = EXCLUDED.rule_a,
                        rule_b = EXCLUDED.rule_b,
                        start_date = EXCLUDED.start_date,
                        end_date = EXCLUDED.end_date,
                        status = EXCLUDED.status,
                        metrics = EXCLUDED.metrics
                    """,
                    (
                        test.id,
                        test.name,
                        json.dumps({"id": test.rule_a.id, "name": test.rule_a.name}),
                        json.dumps({"id": test.rule_b.id, "name": test.rule_b.name}),
                        test.start_date,
                        test.end_date,
                        test.status,
                        json.dumps(test.metrics),
                    ),
                )


class PgTaskRepository:
    def __init__(self):
        self._ensure_table()

    def _ensure_table(self):
        with get_pg_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS tasks (
                        id TEXT PRIMARY KEY,
                        title TEXT NOT NULL,
                        description TEXT,
                        priority TEXT,
                        status TEXT,
                        assigned_to TEXT,
                        due_date TIMESTAMP,
                        created_at TIMESTAMP,
                        entity_type TEXT,
                        entity_id TEXT,
                        task_type TEXT,
                        automation_level TEXT,
                        estimated_hours DOUBLE PRECISION,
                        actual_hours DOUBLE PRECISION,
                        completion_percentage DOUBLE PRECISION,
                        dependencies JSONB,
                        tags JSONB
                    )
                    """
                )

    async def save_task(self, task: Task):
        with get_pg_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO tasks (
                        id, title, description, priority, status, assigned_to, due_date, created_at, entity_type, entity_id,
                        task_type, automation_level, estimated_hours, actual_hours, completion_percentage, dependencies, tags
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (id) DO UPDATE SET
                        title = EXCLUDED.title,
                        description = EXCLUDED.description,
                        priority = EXCLUDED.priority,
                        status = EXCLUDED.status,
                        assigned_to = EXCLUDED.assigned_to,
                        due_date = EXCLUDED.due_date,
                        created_at = EXCLUDED.created_at,
                        entity_type = EXCLUDED.entity_type,
                        entity_id = EXCLUDED.entity_id,
                        task_type = EXCLUDED.task_type,
                        automation_level = EXCLUDED.automation_level,
                        estimated_hours = EXCLUDED.estimated_hours,
                        actual_hours = EXCLUDED.actual_hours,
                        completion_percentage = EXCLUDED.completion_percentage,
                        dependencies = EXCLUDED.dependencies,
                        tags = EXCLUDED.tags
                    """,
                    (
                        task.id,
                        task.title,
                        task.description,
                        task.priority,
                        task.status,
                        task.assigned_to,
                        task.due_date,
                        task.created_at,
                        task.entity_type,
                        task.entity_id,
                        getattr(task, 'task_type', 'Manual'),
                        getattr(task, 'automation_level', 'None'),
                        getattr(task, 'estimated_hours', 0.0),
                        getattr(task, 'actual_hours', 0.0),
                        getattr(task, 'completion_percentage', 0.0),
                        json.dumps(getattr(task, 'dependencies', [])),
                        json.dumps(getattr(task, 'tags', [])),
                    ),
                )

    async def get_tasks(self, status: str = None, priority: str = None) -> List[Dict]:
        with get_pg_connection() as conn:
            with conn.cursor() as cur:
                query = "SELECT * FROM tasks WHERE 1=1"
                params: list = []
                if status:
                    query += " AND status = %s"
                    params.append(status)
                if priority:
                    query += " AND priority = %s"
                    params.append(priority)
                query += " ORDER BY created_at DESC"
                cur.execute(query, params)
                columns = [desc.name for desc in cur.description]
                rows = cur.fetchall()
                return [dict(zip(columns, row)) for row in rows]

    async def update_task_status(self, task_id: str, status: str, completion_percentage: float = None):
        with get_pg_connection() as conn:
            with conn.cursor() as cur:
                if completion_percentage is not None:
                    cur.execute(
                        "UPDATE tasks SET status = %s, completion_percentage = %s WHERE id = %s",
                        (status, completion_percentage, task_id),
                    )
                else:
                    cur.execute(
                        "UPDATE tasks SET status = %s WHERE id = %s",
                        (status, task_id),
                    )

    async def get_overdue_tasks(self) -> List[Dict]:
        with get_pg_connection() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    SELECT * FROM tasks
                    WHERE due_date < NOW() AND status != 'Completed'
                    ORDER BY due_date ASC
                    """
                )
                columns = [desc.name for desc in cur.description]
                rows = cur.fetchall()
                return [dict(zip(columns, row)) for row in rows]
